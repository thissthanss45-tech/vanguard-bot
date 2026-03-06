import json
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import yfinance as yf


FORECAST_LOG_PATH = Path("data/forecast_snapshots.jsonl")


@dataclass
class ForecastSnapshot:
    symbol: str
    analyzed_at_utc: str
    base_price: float
    atr_14: float
    d1_bull: int
    d1_bear: int
    d1_bias: str
    d2_bull: int
    d2_bear: int
    d2_bias: str
    d3_bull: int
    d3_bear: int
    d3_bias: str
    # ML fields (optional, default = not available)
    ml_bull: int = 50
    ml_bear: int = 50
    ml_accuracy_wf: float = 0.0
    ml_confidence: str = "н/д"
    ml_available: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "symbol": self.symbol,
            "analyzed_at_utc": self.analyzed_at_utc,
            "base_price": self.base_price,
            "atr_14": self.atr_14,
            "d1_bull": self.d1_bull,
            "d1_bear": self.d1_bear,
            "d1_bias": self.d1_bias,
            "d2_bull": self.d2_bull,
            "d2_bear": self.d2_bear,
            "d2_bias": self.d2_bias,
            "d3_bull": self.d3_bull,
            "d3_bear": self.d3_bear,
            "d3_bias": self.d3_bias,
            "ml_bull": self.ml_bull,
            "ml_bear": self.ml_bear,
            "ml_accuracy_wf": self.ml_accuracy_wf,
            "ml_confidence": self.ml_confidence,
            "ml_available": self.ml_available,
        }


def _safe_int(value: Any, default: int = 50) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _parse_dt(dt_text: str) -> datetime:
    parsed = datetime.fromisoformat(dt_text.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def append_snapshot_from_market_data(market_data: dict):
    forecast = market_data.get("forecast_3d", [])
    if len(forecast) < 3:
        return

    analyzed_at = datetime.now(timezone.utc).isoformat()
    item1, item2, item3 = forecast[0], forecast[1], forecast[2]
    ml = market_data.get("ml_forecast", {})

    snapshot = ForecastSnapshot(
        symbol=str(market_data.get("symbol", "")).strip().upper(),
        analyzed_at_utc=analyzed_at,
        base_price=float(market_data.get("current_price", 0.0) or 0.0),
        atr_14=float(market_data.get("atr_14", 0.0) or 0.0),
        d1_bull=_safe_int(item1.get("bullish_probability"), 50),
        d1_bear=_safe_int(item1.get("bearish_probability"), 50),
        d1_bias=str(item1.get("bias", "Нейтральный")),
        d2_bull=_safe_int(item2.get("bullish_probability"), 50),
        d2_bear=_safe_int(item2.get("bearish_probability"), 50),
        d2_bias=str(item2.get("bias", "Нейтральный")),
        d3_bull=_safe_int(item3.get("bullish_probability"), 50),
        d3_bear=_safe_int(item3.get("bearish_probability"), 50),
        d3_bias=str(item3.get("bias", "Нейтральный")),
        ml_bull=_safe_int(ml.get("ml_bull_prob"), 50),
        ml_bear=_safe_int(ml.get("ml_bear_prob"), 50),
        ml_accuracy_wf=float(ml.get("ml_accuracy_wf") or 0.0),
        ml_confidence=str(ml.get("ml_confidence") or "н/д"),
        ml_available=bool(ml.get("ml_available", False)),
    )

    FORECAST_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with FORECAST_LOG_PATH.open("a", encoding="utf-8") as fp:
        fp.write(json.dumps(snapshot.to_dict(), ensure_ascii=False) + "\n")


def _load_snapshots() -> list[dict[str, Any]]:
    if not FORECAST_LOG_PATH.exists():
        return []
    rows = []
    with FORECAST_LOG_PATH.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue
    return rows


def _latest_per_symbol(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for row in rows:
        symbol = str(row.get("symbol", "")).strip().upper()
        if not symbol:
            continue
        ts = str(row.get("analyzed_at_utc", ""))
        prev = latest.get(symbol)
        if prev is None or ts > str(prev.get("analyzed_at_utc", "")):
            latest[symbol] = row
    return list(latest.values())


def _is_matured(analyzed_at_utc: str, hours: int = 72) -> bool:
    try:
        dt = _parse_dt(analyzed_at_utc)
    except Exception:
        return False
    return datetime.now(timezone.utc) - dt >= timedelta(hours=hours)


def _eta_to_maturity(analyzed_at_utc: str, hours: int = 72) -> tuple[int | None, str | None]:
    try:
        dt = _parse_dt(analyzed_at_utc)
    except Exception:
        return None, None

    target = dt + timedelta(hours=hours)
    remaining = int((target - datetime.now(timezone.utc)).total_seconds())
    if remaining <= 0:
        return 0, "уже доступен"

    rem_hours = remaining // 3600
    rem_minutes = (remaining % 3600) // 60
    return remaining, f"{rem_hours} ч {rem_minutes} мин"


def _classify_outcome(return_pct: float, atr_14: float, base_price: float) -> str:
    if base_price <= 0:
        return "Нейтральный"

    atr_part = (atr_14 / base_price) * 100 if atr_14 > 0 else 0.5
    threshold = max(0.25, min(1.5, 0.25 * atr_part))

    if return_pct > threshold:
        return "Бычий"
    if return_pct < -threshold:
        return "Медвежий"
    return "Нейтральный"


def _actual_return_3d(symbol: str, analyzed_at_utc: str, base_price: float) -> tuple[float | None, str | None]:
    try:
        start = _parse_dt(analyzed_at_utc)
    except Exception:
        return None, None

    # Compare dates only — ignore time-of-day so a market open at 14:30 UTC
    # on the 3rd calendar day still counts even if snapshot was taken at 02:00 UTC.
    target_date = (start + timedelta(days=3)).date()

    try:
        hist = yf.Ticker(symbol).history(period="14d", interval="1d")
    except Exception:
        return None, None

    if hist is None or hist.empty:
        return None, None

    index_utc = pd.to_datetime(hist.index, utc=True)
    hist = hist.copy()
    hist.index = index_utc
    # Filter rows whose DATE is >= target_date
    eligible = hist[hist.index.normalize() >= pd.Timestamp(target_date, tz="UTC")]
    if eligible.empty:
        return None, None

    close_price = float(eligible["Close"].iloc[0])
    if base_price <= 0:
        return None, None

    ret = ((close_price - base_price) / base_price) * 100
    return round(ret, 2), close_price.__str__()


def build_matured_report(max_items: int = 20) -> tuple[str, list[dict[str, Any]]]:
    # Use ALL snapshots (not just latest per symbol) for a complete picture
    all_rows = _load_snapshots()
    latest_per_sym = _latest_per_symbol(all_rows)  # for pending ETA only
    matured = [row for row in all_rows if _is_matured(str(row.get("analyzed_at_utc", "")), 72)]
    matured = sorted(matured, key=lambda row: row.get("analyzed_at_utc", ""), reverse=True)

    if not matured:
        pending = []
        for row in latest_per_sym:
            eta_seconds, eta_text = _eta_to_maturity(str(row.get("analyzed_at_utc", "")), 72)
            if eta_seconds is None:
                continue
            if eta_seconds > 0:
                pending.append((eta_seconds, str(row.get("symbol", "?")), eta_text))

        if not pending:
            return "Пока нет тикеров с созревшим горизонтом D3 (3+ дня).", []

        pending.sort(key=lambda item: item[0])
        _, symbol, eta_text = pending[0]

        recent_rows = sorted(
            [row for row in latest_per_sym if not _is_matured(str(row.get("analyzed_at_utc", "")), 72)],
            key=lambda row: str(row.get("analyzed_at_utc", "")),
            reverse=True,
        )

        lines = [
            "Пока нет тикеров с созревшим горизонтом D3 (3+ дня).",
            f"Ближайший: {symbol} примерно через {eta_text}.",
        ]

        if recent_rows:
            lines.append("Последние ожидающие тикеры:")
            for row in recent_rows[:5]:
                s = str(row.get("symbol", "?"))
                _, eta_item = _eta_to_maturity(str(row.get("analyzed_at_utc", "")), 72)
                lines.append(f"• {s}: примерно через {eta_item or 'н/д'}")

        return "\n".join(lines), []

    rows: list[dict[str, Any]] = []
    lines = ["📈 Анализ по тикерам (созревшие D3):"]

    for item in matured[:max_items]:
        symbol = str(item.get("symbol", ""))
        base_price = float(item.get("base_price", 0.0) or 0.0)
        atr_14 = float(item.get("atr_14", 0.0) or 0.0)
        d3_bias = str(item.get("d3_bias", "Нейтральный"))
        d3_bull = _safe_int(item.get("d3_bull"), 50)
        d3_bear = _safe_int(item.get("d3_bear"), 50)

        ret_3d, close_text = _actual_return_3d(symbol, str(item.get("analyzed_at_utc", "")), base_price)
        if ret_3d is None:
            actual_bias = "н/д"
            ret_text = "н/д"
            verdict = "pending"
        else:
            actual_bias = _classify_outcome(ret_3d, atr_14, base_price)
            ret_text = f"{ret_3d:+.2f}%"
            verdict = "✅" if actual_bias == d3_bias else "🔴"

        lines.append(
            f"• {symbol}: прогноз={d3_bias}({d3_bull}%) | факт={actual_bias} | доходность={ret_text} | {verdict}"
        )

        rows.append(
            {
                "symbol": symbol,
                "analyzed_at_utc": item.get("analyzed_at_utc"),
                "base_price": round(base_price, 4),
                "d3_bias_forecast": d3_bias,
                "d3_bull_forecast": d3_bull,
                "d3_bear_forecast": d3_bear,
                "actual_return_3d_pct": ret_3d,
                "actual_bias_3d": actual_bias,
                "verdict": verdict,
                "close_used": close_text,
                "ml_bull": item.get("ml_bull", 50),
                "ml_accuracy_wf": item.get("ml_accuracy_wf", 0.0),
                "ml_available": item.get("ml_available", False),
            }
        )

    return "\n".join(lines), rows


def build_accuracy_stats() -> str:
    """
    Aggregate D3-forecast accuracy over ALL matured snapshots.
    Returns a human-readable stats report with:
      - overall correct/incorrect/total counts + accuracy %
      - accuracy breakdown by bias (bullish/bearish/neutral)
      - top-5 symbols ranked by most forecasts
      - ML vs rule-only comparison (when ML was available)
    """
    all_rows = _load_snapshots()
    matured = [row for row in all_rows if _is_matured(str(row.get("analyzed_at_utc", "")), 72)]

    if not matured:
        return "📊 Нет созревших данных для анализа точности (нужно 3+ дня после первого прогноза)."

    total = 0
    correct = 0
    incorrect = 0
    pending_count = 0

    by_bias: dict[str, dict[str, int]] = {}
    by_symbol: dict[str, int] = {}

    ml_total = 0
    ml_correct = 0
    rule_only_total = 0
    rule_only_correct = 0

    for item in matured:
        symbol = str(item.get("symbol", "?"))
        base_price = float(item.get("base_price", 0.0) or 0.0)
        atr_14 = float(item.get("atr_14", 0.0) or 0.0)
        d3_bias = str(item.get("d3_bias", "Нейтральный"))

        ret_3d, _ = _actual_return_3d(symbol, str(item.get("analyzed_at_utc", "")), base_price)
        if ret_3d is None:
            pending_count += 1
            continue

        actual_bias = _classify_outcome(ret_3d, atr_14, base_price)
        is_correct = actual_bias == d3_bias
        total += 1
        if is_correct:
            correct += 1
        else:
            incorrect += 1

        # by bias bucket
        bucket = by_bias.setdefault(d3_bias, {"correct": 0, "total": 0})
        bucket["total"] += 1
        if is_correct:
            bucket["correct"] += 1

        # by symbol
        by_symbol[symbol] = by_symbol.get(symbol, 0) + 1

        # ML vs rule-only
        ml_avail = bool(item.get("ml_available", False))
        if ml_avail:
            ml_total += 1
            if is_correct:
                ml_correct += 1
        else:
            rule_only_total += 1
            if is_correct:
                rule_only_correct += 1

    if total == 0:
        return (
            "📊 Все снапшоты ещё ожидают созревания или нет данных с yfinance.\n"
            f"Ожидают: {pending_count} прогнозов."
        )

    accuracy = correct / total * 100
    lines = [
        "📊 Статистика точности D3-прогнозов",
        f"Всего оценено: {total} | ✅ Верных: {correct} | 🔴 Ошибок: {incorrect}",
        f"Точность (D3 bias): {accuracy:.1f}%",
    ]

    if pending_count:
        lines.append(f"⌛ Ещё ожидают созревания: {pending_count}")

    # Breakdown by bias
    lines.append("")
    lines.append("По типу прогноза:")
    for bias_name in ["Бычий", "Медвежий", "Нейтральный"]:
        b = by_bias.get(bias_name)
        if b and b["total"] > 0:
            b_acc = b["correct"] / b["total"] * 100
            lines.append(f"  • {bias_name}: {b['correct']}/{b['total']} ({b_acc:.0f}%)")

    # ML vs rule-only
    if ml_total > 0 and rule_only_total > 0:
        lines.append("")
        lines.append("ML-ансамбль vs только правила:")
        lines.append(f"  • С ML: {ml_correct}/{ml_total} ({ml_correct / ml_total * 100:.0f}%)")
        lines.append(f"  • Только правила: {rule_only_correct}/{rule_only_total} ({rule_only_correct / rule_only_total * 100:.0f}%)")
    elif ml_total > 0:
        lines.append(f"\nС ML-ансамблем: {ml_correct}/{ml_total} ({ml_correct / ml_total * 100:.0f}%)")

    # Top-5 symbols by forecast count
    top5 = sorted(by_symbol.items(), key=lambda kv: kv[1], reverse=True)[:5]
    if top5:
        lines.append("")
        lines.append("Топ тикеров по количеству прогнозов:")
        for sym, cnt in top5:
            lines.append(f"  • {sym}: {cnt}")

    lines.append("")
    lines.append("ℹ️ Данные из data/forecast_snapshots.jsonl")
    return "\n".join(lines)


def build_per_ticker_accuracy(min_forecasts: int = 1) -> str:
    """
    Per-ticker D3 accuracy breakdown.
    Columns: ticker | forecasts | correct | accuracy% | avg_return% | last bias
    Sorted by accuracy desc, then by forecast count desc.
    Only tickers with >= min_forecasts evaluated results are shown.
    """
    all_rows = _load_snapshots()
    matured = [row for row in all_rows if _is_matured(str(row.get("analyzed_at_utc", "")), 72)]

    if not matured:
        return "📊 Нет созревших данных (нужно 3+ дня после первого прогноза)."

    # symbol → {total, correct, pending, returns[], last_bias, last_ts}
    stats: dict[str, dict] = {}

    for item in matured:
        symbol = str(item.get("symbol", "?"))
        base_price = float(item.get("base_price", 0.0) or 0.0)
        atr_14 = float(item.get("atr_14", 0.0) or 0.0)
        d3_bias = str(item.get("d3_bias", "Нейтральный"))
        analyzed_at = str(item.get("analyzed_at_utc", ""))

        s = stats.setdefault(symbol, {
            "total": 0, "correct": 0, "pending": 0,
            "returns": [], "last_bias": d3_bias, "last_ts": analyzed_at,
        })

        # Keep track of latest snapshot
        if analyzed_at > s["last_ts"]:
            s["last_ts"] = analyzed_at
            s["last_bias"] = d3_bias

        ret_3d, _ = _actual_return_3d(symbol, analyzed_at, base_price)
        if ret_3d is None:
            s["pending"] += 1
            continue

        actual_bias = _classify_outcome(ret_3d, atr_14, base_price)
        s["total"] += 1
        if actual_bias == d3_bias:
            s["correct"] += 1
        s["returns"].append(ret_3d)

    # Filter tickers with enough data
    eligible = {sym: v for sym, v in stats.items() if v["total"] >= min_forecasts}

    if not eligible:
        return (
            "📊 Недостаточно данных для расчёта per-ticker точности.\n"
            f"Минимум прогнозов на тикер: {min_forecasts}. "
            f"Тикеров в базе: {len(stats)}."
        )

    # Sort: accuracy desc, then total desc
    sorted_tickers = sorted(
        eligible.items(),
        key=lambda kv: (kv[1]["correct"] / kv[1]["total"], kv[1]["total"]),
        reverse=True,
    )

    _BIAS_EMOJI = {"Бычий": "🐂", "Медвежий": "🐻", "Нейтральный": "⚖️"}
    lines = [
        "📈 Точность прогнозов по тикерам (D3)\n",
        f"{'Тикер':<10} {'Σ':>4} {'✅':>4} {'%':>6} {'Δ avg':>8}  Последний",
        "─" * 46,
    ]

    for sym, v in sorted_tickers:
        total = v["total"]
        correct = v["correct"]
        acc = correct / total * 100
        avg_ret = sum(v["returns"]) / len(v["returns"]) if v["returns"] else 0.0
        last_bias = v["last_bias"]
        emoji = _BIAS_EMOJI.get(last_bias, "⚖️")
        # pending note
        pend_note = f" (+{v['pending']}⌛)" if v["pending"] else ""
        lines.append(
            f"{sym:<10} {total:>4}  {correct:>3}  {acc:>5.0f}%  {avg_ret:>+6.1f}%  {emoji}{last_bias}{pend_note}"
        )

    overall_total = sum(v["total"] for v in eligible.values())
    overall_correct = sum(v["correct"] for v in eligible.values())
    pending_total = sum(v["pending"] for v in stats.values())
    overall_acc = overall_correct / overall_total * 100 if overall_total else 0

    lines.append("─" * 46)
    lines.append(
        f"{'ИТОГО':<10} {overall_total:>4}  {overall_correct:>3}  {overall_acc:>5.0f}%"
    )
    if pending_total:
        lines.append(f"⌛ Ещё ожидают созревания: {pending_total}")
    lines.append("\nℹ️ Данные из data/forecast_snapshots.jsonl")
    return "\n".join(lines)


def build_ticker_backtest(ticker: str) -> str:
    """
    Детальный бэктест по одному тикеру.
    Показывает таблицу всех созревших прогнозов: дата, bias, факт, результат, доходность.
    """
    ticker = ticker.upper()
    all_rows = _load_snapshots()
    rows = [r for r in all_rows if str(r.get("symbol", "")).upper() == ticker]

    if not rows:
        return f"📊 Нет данных для {ticker}. Сначала запусти AI-Анализ по этому тикеру."

    matured = [r for r in rows if _is_matured(str(r.get("analyzed_at_utc", "")), 72)]
    pending = [r for r in rows if not _is_matured(str(r.get("analyzed_at_utc", "")), 72)]

    if not matured:
        return (
            f"📊 {ticker}: {len(pending)} прогнозов ожидают созревания (нужно 3+ дней).\n"
            "Первый прогноз:\n" + str(rows[0].get("analyzed_at_utc", "?"))
        )

    _BIAS_EMOJI = {"Бычий": "🐂", "Медвежий": "🐻", "Нейтральный": "⚖️"}
    correct = 0
    returns = []
    detail_lines = []

    for item in matured:
        base_price = float(item.get("base_price", 0) or 0)
        atr_14 = float(item.get("atr_14", 0) or 0)
        d3_bias = str(item.get("d3_bias", "Нейтральный"))
        analyzed_at = str(item.get("analyzed_at_utc", ""))
        date_short = analyzed_at[:10]

        ret_3d, _ = _actual_return_3d(ticker, analyzed_at, base_price)
        if ret_3d is None:
            continue

        actual_bias = _classify_outcome(ret_3d, atr_14, base_price)
        is_correct = actual_bias == d3_bias
        if is_correct:
            correct += 1
        returns.append(ret_3d)

        verdict = "✅" if is_correct else "🔴"
        pred_e = _BIAS_EMOJI.get(d3_bias, "⚖️")
        actual_e = _BIAS_EMOJI.get(actual_bias, "⚖️")
        ret_str = f"{ret_3d:+.1f}%" if ret_3d is not None else "н/д"
        detail_lines.append(f"{date_short} {pred_e}{d3_bias[:3]:>3} → {actual_e}{actual_bias[:3]:>3} {ret_str} {verdict}")

    total = len(returns)
    acc = correct / total * 100 if total else 0
    avg_ret = sum(returns) / total if returns else 0
    pos_trades = sum(1 for r in returns if r > 0)
    win_rate = pos_trades / total * 100 if total else 0

    header = (
        f"📊 БЭКТЕСТ {ticker} (D3 прогнозы)\n"
        f"{'─'*36}\n"
        f"Прогнозов: {total}  |  Точность: {acc:.0f}%\n"
        f"Win Rate (по доходности): {win_rate:.0f}%\n"
        f"Средняя доходность: {avg_ret:+.2f}%\n"
    )
    if pending:
        header += f"⌛ Ещё в ожидании: {len(pending)}\n"
    header += f"{'─'*36}\n"
    detail = "\n".join(detail_lines[-20:])  # последние 20
    footer = "\n\nℹ️ Показаны последние 20 созревших прогнозов"

    return header + detail + (footer if len(detail_lines) > 20 else "")


def purge_old_snapshots(keep_days: int = 7) -> tuple[int, int]:
    """
    Удаляет снапшоты старше keep_days дней.
    Возвращает (удалено, осталось).
    """
    all_rows = _load_snapshots()
    if not all_rows:
        return 0, 0

    cutoff = datetime.now(tz=timezone.utc) - timedelta(days=keep_days)
    kept = []
    removed = 0
    for row in all_rows:
        ts_str = str(row.get("analyzed_at_utc", ""))
        try:
            ts = datetime.fromisoformat(ts_str)
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
        except ValueError:
            kept.append(row)  # не можем распарсить — оставляем
            continue
        if ts >= cutoff:
            kept.append(row)
        else:
            removed += 1

    FORECAST_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(FORECAST_LOG_PATH, "w", encoding="utf-8") as fh:
        for row in kept:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")

    return removed, len(kept)


def export_matured_report_to_excel(out_path: str = "data/forecast_report_3d.xlsx") -> str | None:
    _, rows = build_matured_report(max_items=500)
    if not rows:
        return None

    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Column order and Russian headers
    col_map = {
        "symbol": "Тикер",
        "analyzed_at_utc": "Дата прогноза (UTC)",
        "base_price": "Цена на момент прогноза",
        "d3_bias_forecast": "Прогноз (bias)",
        "d3_bull_forecast": "Вероятность Бычий, %",
        "d3_bear_forecast": "Вероятность Медвежий, %",
        "actual_return_3d_pct": "Факт. доходность за 3д, %",
        "actual_bias_3d": "Факт. направление",
        "close_used": "Цена закрытия (факт)",
        "verdict": "Итог",
        "ml_bull": "ML вероятность Бычий, %",
        "ml_accuracy_wf": "ML точность WF, %",
        "ml_available": "ML использовался",
    }

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "D3 Прогнозы"

    # Headers
    headers = [col_map.get(k, k) for k in col_map]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        verdict = row_data.get("verdict", "")
        fill = GREEN_FILL if verdict == "✅" else (RED_FILL if verdict == "🔴" else None)
        for col_idx, key in enumerate(col_map, 1):
            val = row_data.get(key, "")
            # Convert bool to readable
            if isinstance(val, bool):
                val = "Да" if val else "Нет"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(header), 10)
        for row_idx in range(2, len(rows) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 35)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Сводка")
    total = len(rows)
    correct = sum(1 for r in rows if r.get("verdict") == "✅")
    incorrect = sum(1 for r in rows if r.get("verdict") == "🔴")
    accuracy = correct / total * 100 if total else 0
    summary_rows = [
        ("Всего прогнозов", total),
        ("✅ Верных", correct),
        ("🔴 Ошибочных", incorrect),
        ("Точность, %", round(accuracy, 1)),
    ]
    for i, (label, val) in enumerate(summary_rows, 1):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=val)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 12

    # ── Per-ticker accuracy sheet ──────────────────────────────────────
    ws3 = wb.create_sheet("По тикерам")
    pt_headers = ["Тикер", "Прогнозов", "✅ Верных", "Точность, %", "Ср. доходность, %", "Ожидают"]
    for col_idx, hdr in enumerate(pt_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Build per-ticker from rows list
    pt_stats: dict[str, dict] = {}
    for r in rows:
        sym = str(r.get("symbol", "?"))
        s = pt_stats.setdefault(sym, {"total": 0, "correct": 0, "returns": []})
        s["total"] += 1
        if r.get("verdict") == "✅":
            s["correct"] += 1
        ret = r.get("actual_return_3d_pct")
        if ret is not None:
            s["returns"].append(float(ret))

    # Count pending from full snapshots for completeness
    all_snap = _load_snapshots()
    pending_by_sym: dict[str, int] = {}
    for snap in all_snap:
        if not _is_matured(str(snap.get("analyzed_at_utc", "")), 72):
            sym = str(snap.get("symbol", "?"))
            pending_by_sym[sym] = pending_by_sym.get(sym, 0) + 1

    pt_sorted = sorted(
        pt_stats.items(),
        key=lambda kv: (kv[1]["correct"] / kv[1]["total"] if kv[1]["total"] else 0, kv[1]["total"]),
        reverse=True,
    )
    for row_idx, (sym, v) in enumerate(pt_sorted, 2):
        t = v["total"]
        c = v["correct"]
        acc_val = round(c / t * 100, 1) if t else 0
        avg_ret = round(sum(v["returns"]) / len(v["returns"]), 2) if v["returns"] else 0
        pend = pending_by_sym.get(sym, 0)
        row_fill = GREEN_FILL if acc_val >= 60 else (RED_FILL if acc_val < 40 else None)
        for col_idx, val in enumerate([sym, t, c, acc_val, avg_ret, pend], 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if row_fill:
                cell.fill = row_fill

    for col_idx in range(1, len(pt_headers) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 18
    ws3.freeze_panes = "A2"

    wb.save(str(path))
    return str(path)
